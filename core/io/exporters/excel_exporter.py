#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Excel Exporter
===============
Export analysis results to formatted Excel workbook with 4 sheets.
"""

import io
from typing import Optional, Dict, Any, List
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_excel_workbook(
    basic_data: Optional[Dict[str, Any]] = None,
    dose_response_data: Optional[Dict[str, Any]] = None,
    thermodynamics_data: Optional[Dict[str, Any]] = None,
    settings_data: Optional[Dict[str, Any]] = None
) -> bytes:
    """
    Create a formatted Excel workbook with 4 sheets containing all analysis results.

    Args:
        basic_data: Basic analysis results from analysis-results-store
        dose_response_data: EC50 analysis results from dose-response-store
        thermodynamics_data: Van't Hoff results from thermodynamics-store
        settings_data: Analysis settings and metadata

    Returns:
        Excel file bytes

    Sheet Structure:
        1. Basic_Analysis - Tm results table
        2. Dose_Response - EC50 analysis table
        3. Thermodynamics - Van't Hoff parameters
        4. Analysis_Settings - All user settings and QC thresholds
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Basic Analysis
        _write_basic_analysis_sheet(writer, basic_data)

        # Sheet 2: Dose Response
        _write_dose_response_sheet(writer, dose_response_data)

        # Sheet 3: Thermodynamics
        _write_thermodynamics_sheet(writer, thermodynamics_data)

        # Sheet 4: Analysis Settings
        _write_settings_sheet(writer, settings_data, basic_data)

    # Apply formatting
    output.seek(0)
    wb = load_workbook(output)
    _apply_excel_formatting(wb)

    # Save to new buffer
    formatted_output = io.BytesIO()
    wb.save(formatted_output)
    formatted_output.seek(0)

    return formatted_output.getvalue()


def _write_basic_analysis_sheet(writer: pd.ExcelWriter, data: Optional[Dict[str, Any]]):
    """Write Basic Analysis sheet with Tm results."""
    if data and 'results' in data and len(data['results']) > 0:
        results = data['results']

        # Build DataFrame
        rows = []
        for r in results:
            row = {
                'Sample': r.get('name', ''),
                'Concentration (M)': r.get('concentration'),
                'Tm (°C)': r.get('tm'),
                'Tm Error (°C)': r.get('tm_error'),
                'R²': r.get('r_squared'),
                'Method': r.get('method', '').upper(),
                'QC Status': r.get('quality_status', ''),
                'QC Flag': r.get('quality_flag', ''),
                'Source File': r.get('source_file', data.get('filenames', [''])[0] if data.get('filenames') else '')
            }
            rows.append(row)

        df = pd.DataFrame(rows)
    else:
        # Empty placeholder
        df = pd.DataFrame({
            'Sample': [],
            'Concentration (M)': [],
            'Tm (°C)': [],
            'Tm Error (°C)': [],
            'R²': [],
            'Method': [],
            'QC Status': [],
            'QC Flag': [],
            'Source File': []
        })
        # Add note row
        df.loc[0] = ['No Basic Analysis results. Please run analysis first.', '', '', '', '', '', '', '', '']

    df.to_excel(writer, sheet_name='Basic_Analysis', index=False)


def _write_dose_response_sheet(writer: pd.ExcelWriter, data: Optional[Dict[str, Any]]):
    """Write Dose Response sheet with EC50 results and SFQ analysis."""
    if data and data.get('ec50') is not None:
        # Extract SFQ data if available
        sfq = data.get('sfq_result', {}) or {}

        # Single row table with dose-response parameters + SFQ
        rows = [{
            'EC50 (M)': data.get('ec50'),
            'EC50 CI Lower (M)': data['ec50_ci'][0] if data.get('ec50_ci') else None,
            'EC50 CI Upper (M)': data['ec50_ci'][1] if data.get('ec50_ci') else None,
            'R²': data.get('r2'),
            'Hill Slope': data.get('hill_slope'),
            'Bottom (°C)': data.get('bottom'),
            'Top (°C)': data.get('top'),
            'N Points': data.get('n_points'),
            'QC Flag': data.get('qc_flag', ''),
            'QC Score': data.get('qc_score'),
            'QC Message': data.get('qc_message', ''),
            'QC Details': data.get('qc_tooltip', ''),
            # SFQ fields
            'SFQ Status': sfq.get('dataset_status', 'N/A'),
            'SFQ Channel': sfq.get('channel_name', ''),
            'SFQ Mode': sfq.get('mode', ''),
            'SFQ EC50_app': sfq.get('ec50_app_str', ''),
            'SFQ Span (%)': sfq.get('span'),
            'SFQ ΔAIC': sfq.get('delta_aic'),
            'SFQ SI': sfq.get('saturation_index'),
            'SFQ Notes': sfq.get('notes', ''),
        }]
        df = pd.DataFrame(rows)
    else:
        # Empty placeholder
        df = pd.DataFrame({
            'EC50 (M)': [],
            'EC50 CI Lower (M)': [],
            'EC50 CI Upper (M)': [],
            'R²': [],
            'Hill Slope': [],
            'Bottom (°C)': [],
            'Top (°C)': [],
            'N Points': [],
            'QC Flag': [],
            'QC Score': [],
            'QC Message': [],
            'QC Details': [],
            'SFQ Status': [],
            'SFQ Channel': [],
            'SFQ Mode': [],
            'SFQ EC50_app': [],
            'SFQ Span (%)': [],
            'SFQ ΔAIC': [],
            'SFQ SI': [],
            'SFQ Notes': []
        })
        df.loc[0] = ['No Dose-Response analysis run. Please navigate to Dose-Response tab and run analysis to generate EC50 data.',
                     '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '']

    df.to_excel(writer, sheet_name='Dose_Response', index=False)


def _write_thermodynamics_sheet(writer: pd.ExcelWriter, data: Optional[Dict[str, Any]]):
    """Write Thermodynamics sheet with Van't Hoff parameters."""
    if data and data.get('delta_h') is not None:
        # Parameter table format
        delta_h_unit = data.get('delta_h_unit', 'kJ/mol')
        delta_s_unit = data.get('delta_s_unit', 'cal/mol·K')

        # Format KD values
        kd_298_val = data.get('kd_298k')
        kd_310_val = data.get('kd_310k')

        # Convert to nM for display
        kd_298_nm = f"{kd_298_val * 1e9:.1f}" if kd_298_val else "N/A"
        kd_310_nm = f"{kd_310_val * 1e9:.1f}" if kd_310_val else "N/A"

        rows = [
            {'Parameter': 'R²', 'Value': data.get('r2'), 'Unit': '-', 'QC Status': data.get('qc_flag', '')},
            {'Parameter': 'QC Score', 'Value': data.get('qc_score'), 'Unit': '/100', 'QC Status': ''},
            {'Parameter': 'N Points', 'Value': data.get('n_points'), 'Unit': '-', 'QC Status': ''},
            {'Parameter': 'ΔH', 'Value': data.get('delta_h_display'), 'Unit': delta_h_unit, 'QC Status': ''},
            {'Parameter': 'ΔS', 'Value': data.get('delta_s_display'), 'Unit': delta_s_unit, 'QC Status': ''},
            {'Parameter': 'KD (298K / 25°C)', 'Value': kd_298_nm, 'Unit': 'nM', 'QC Status': ''},
            {'Parameter': 'KD (310K / 37°C)', 'Value': kd_310_nm, 'Unit': 'nM', 'QC Status': ''},
        ]
        df = pd.DataFrame(rows)

        # Add QC message and details
        if data.get('qc_message'):
            note_row = pd.DataFrame([{
                'Parameter': 'QC Summary',
                'Value': data['qc_message'],
                'Unit': '',
                'QC Status': ''
            }])
            df = pd.concat([df, note_row], ignore_index=True)

        # Add QC details if available
        if data.get('qc_details'):
            qc_details = data['qc_details']
            if isinstance(qc_details, dict):
                # Add each detail as a separate row
                for key, value in qc_details.items():
                    detail_row = pd.DataFrame([{
                        'Parameter': f'  QC Detail: {key}',
                        'Value': str(value),
                        'Unit': '',
                        'QC Status': ''
                    }])
                    df = pd.concat([df, detail_row], ignore_index=True)
    else:
        # Empty placeholder
        df = pd.DataFrame({
            'Parameter': ['No Thermodynamics analysis run. Please navigate to Thermodynamics tab and run Van\'t Hoff analysis to generate thermodynamic parameters.'],
            'Value': [''],
            'Unit': [''],
            'QC Status': ['']
        })

    df.to_excel(writer, sheet_name='Thermodynamics', index=False)


def _write_settings_sheet(writer: pd.ExcelWriter, settings_data: Optional[Dict[str, Any]], basic_data: Optional[Dict[str, Any]]):
    """Write Analysis Settings sheet with all parameters and metadata."""
    rows = []

    # Section 1: Basic Analysis Settings
    rows.append({'Section': 'Basic Analysis Settings', 'Parameter': '', 'Value': '', 'Description': ''})
    basic_settings = settings_data.get('basic_analysis', {}) if settings_data else {}
    rows.append({'Section': '', 'Parameter': 'Tm Method', 'Value': basic_settings.get('method', 'AUC'), 'Description': 'Method used for Tm calculation'})
    rows.append({'Section': '', 'Parameter': 'Channel', 'Value': basic_settings.get('channel', '350nm'), 'Description': 'Fluorescence channel'})
    rows.append({'Section': '', 'Parameter': 'QC Enabled', 'Value': 'Yes', 'Description': 'Quality control checks active'})

    rows.append({'Section': '', 'Parameter': '', 'Value': '', 'Description': ''})  # Spacer

    # Section 2: Dose-Response Settings
    rows.append({'Section': 'Dose-Response Settings', 'Parameter': '', 'Value': '', 'Description': ''})
    dr_settings = settings_data.get('dose_response', {}) if settings_data else {}
    rows.append({'Section': '', 'Parameter': 'Fitting Method', 'Value': '4-Parameter Logistic', 'Description': 'Hill equation variant'})
    rows.append({'Section': '', 'Parameter': 'QC Enabled', 'Value': 'Yes', 'Description': 'Quality control checks active'})

    rows.append({'Section': '', 'Parameter': '', 'Value': '', 'Description': ''})  # Spacer

    # Section 3: Thermodynamics Settings
    rows.append({'Section': 'Thermodynamics Settings', 'Parameter': '', 'Value': '', 'Description': ''})
    thermo_settings = settings_data.get('thermodynamics', {}) if settings_data else {}
    rows.append({'Section': '', 'Parameter': 'Unit System', 'Value': thermo_settings.get('units', 'Calorie'), 'Description': 'kcal/mol vs kJ/mol'})
    rows.append({'Section': '', 'Parameter': 'Temperature Slices', 'Value': thermo_settings.get('n_slices', 5), 'Description': 'Number of isothermal points'})
    rows.append({'Section': '', 'Parameter': 'QC Enabled', 'Value': 'Yes', 'Description': 'Quality control checks active'})

    rows.append({'Section': '', 'Parameter': '', 'Value': '', 'Description': ''})  # Spacer

    # Section 4: QC Thresholds
    rows.append({'Section': 'QC Thresholds', 'Parameter': '', 'Value': '', 'Description': ''})
    rows.append({'Section': '', 'Parameter': 'Minimum R² (Critical)', 'Value': '0.80', 'Description': 'Tm Analysis'})
    rows.append({'Section': '', 'Parameter': 'Recommended R²', 'Value': '0.95', 'Description': 'Tm Analysis'})
    rows.append({'Section': '', 'Parameter': 'Minimum Data Points', 'Value': '3', 'Description': 'Dose-Response'})
    rows.append({'Section': '', 'Parameter': 'Van\'t Hoff R² (Critical)', 'Value': '0.80', 'Description': 'Thermodynamics'})

    rows.append({'Section': '', 'Parameter': '', 'Value': '', 'Description': ''})  # Spacer

    # Section 5: Export Metadata
    rows.append({'Section': 'Export Metadata', 'Parameter': '', 'Value': '', 'Description': ''})
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rows.append({'Section': '', 'Parameter': 'Export Date', 'Value': now, 'Description': ''})
    rows.append({'Section': '', 'Parameter': 'QuantDSF Version', 'Value': '0.9.0', 'Description': ''})

    # Uploaded files
    if basic_data and basic_data.get('filenames'):
        filenames_str = ', '.join(basic_data['filenames'])
    else:
        filenames_str = 'N/A'
    rows.append({'Section': '', 'Parameter': 'Uploaded Files', 'Value': filenames_str, 'Description': ''})

    df = pd.DataFrame(rows)
    df.to_excel(writer, sheet_name='Analysis_Settings', index=False)


def _apply_excel_formatting(wb):
    """Apply formatting to all sheets in the workbook."""
    # Define styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    border_side = Side(style='thin', color='D3D3D3')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    # QC flag colors
    qc_colors = {
        '✅': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
        '⚠️': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
        '❌': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Apply borders and number formatting to data
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

                # Number formatting
                if cell.column_letter in ['B', 'C', 'D']:  # Concentration, Tm, Tm Error columns
                    if isinstance(cell.value, (int, float)):
                        if cell.column == 2:  # Concentration
                            cell.number_format = '0.00E+00'
                        else:  # Tm, errors
                            cell.number_format = '0.0'

                # R² formatting
                if 'R²' in str(ws.cell(row=1, column=cell.column).value):
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0.000'

                # QC Flag conditional formatting
                if cell.value in qc_colors:
                    cell.fill = qc_colors[cell.value]

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
